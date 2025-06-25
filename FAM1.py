import pandas as pd
import pulp
from openpyxl.styles import Font
from openpyxl import load_workbook

# --- Parameters ---
input_file = "CaiGiza Demand Table.xlsx"
output_file = "drone_assignment_results_CaiGiz.xlsx"
total_drones_available = 75 # Total fleet size
max_waiting_time = 1 # Max allowed wait (hours)
drone_cycle_time = 0.30 # 25 minutes in hours

# --- Load Data ---
df_demand = pd.read_excel(input_file, sheet_name="Distance Matrix")[["Hospital Name", "Daily Demand"]]
print("Demand Data:\n", df_demand)

# --- LP Problem Setup ---
prob = pulp.LpProblem("Drone_Assignment", pulp.LpMinimize)
hospitals = df_demand["Hospital Name"].tolist()

# --- Decision Variables ---
x = pulp.LpVariable.dicts(
    "Drones_Assigned",
    hospitals,
    lowBound=1,  # Minimum 1 drone per hub
    cat="Integer"
)

# --- Objective Function ---
prob += pulp.lpSum([x[i] for i in hospitals]), "Total_Drones_Used"

# --- Constraints ---
# 1. Demand satisfaction with cycle time adjustment
for i in hospitals:
    demand = df_demand.loc[df_demand["Hospital Name"] == i, "Daily Demand"].values[0]
    prob += x[i] >= (demand * drone_cycle_time) / max_waiting_time, f"Demand_Satisfaction_{i}"

# 2. Fleet availability (hard constraint)
prob += pulp.lpSum([x[i] for i in hospitals]) <= total_drones_available, "Fleet_Availability"

# --- Solve ---
prob.solve()
print("Solver Status:", pulp.LpStatus[prob.status])

# --- Results ---
results = []
for i in hospitals:
    results.append({
        "Hospital Name": i,
        "Drones_Assigned": int(pulp.value(x[i])),
        "Deliveries/Hour": round(pulp.value(x[i]) / drone_cycle_time, 2)
    })

df_results = pd.DataFrame(results)
total_drones_assigned = df_results["Drones_Assigned"].sum()

# --- Fleet Check & Message ---
if total_drones_assigned > total_drones_available:
    deficit = total_drones_assigned - total_drones_available
    message = f"❌ Shortage: {deficit} more drones needed (Total: {total_drones_assigned})"
else:
    message = f"✅ Fleet sufficient! Total drones assigned: {total_drones_assigned}/{total_drones_available}"

print("\n" + message + "\n")

# --- Excel Export with Formatting ---
df_results.to_excel(output_file, index=False)

# Add formatted message
wb = load_workbook(output_file)
ws = wb.active
ws.append([])  # Empty row
ws.append(["System Message:", message])
ws.cell(row=ws.max_row, column=2).font = Font(bold=True, color="FF0000" if "❌" in message else "007500")

# Add cycle time note
ws.append([])
ws.append(["Note:", f"Drone cycle time = {drone_cycle_time*60} minutes per delivery"])
wb.save(output_file)

print(f"Results saved to {output_file}")